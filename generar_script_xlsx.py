import openpyxl
from openpyxl.styles import PatternFill, Font
import os

# Configuration
# input_file and output_file are now handled via arguments/logic


# Key columns for WHERE clause in UPDATE
# key_columns = ['NUMERO_INSTITUCION', 'TRANSACCION_EXTERNA']

def format_value(val):
    if val is None:
        return 'NULL'
    s_val = str(val).strip()
    if s_val.upper() == 'NULL':
        return 'NULL'
    if s_val.upper() == 'GETDATE()':
        return 'GETDATE()'
    # Escape single quotes
    return f"'{s_val.replace("'", "''")}'"

def is_cell_styled(cell):
    # Check for Bold
    if cell.font and cell.font.bold:
        return True
    
    # Check for Fill Color (Background)
    # patternType='solid' usually indicates a color is set. 
    # specific colors could be checked if needed, but 'change of style' usually implies any non-default.
    if cell.fill and cell.fill.patternType == 'solid':
        # You might want to exclude white or default if the sheet has them, 
        # but usually 'solid' means a deliberate color.
        if cell.fill.fgColor.rgb != '00000000': # 00000000 is often transparent/default in openpyxl
             return True
             
    return False

import argparse

def main():
    parser = argparse.ArgumentParser(description='Generate SQL scripts from Excel file.')
    parser.add_argument('input_file', help='Path to the input Excel file')
    parser.add_argument('--key-columns', '-k', help="Key columns. Format: 'COL1,COL2' (global) or 'Sheet1:COL1,COL2;Sheet2:COL3;DefaultCOL1' (per sheet)", default='NUMERO_INSTITUCION,TRANSACCION_EXTERNA')
    args = parser.parse_args()

    input_file = args.input_file
    
    # Parse key columns configuration
    sheet_keys_config = {}
    global_default_keys = ['NUMERO_INSTITUCION', 'TRANSACCION_EXTERNA']
    
    # If the argument is exactly the default, treat it as global default without parsing for sheets (simple case)
    # However, our parser handles it fine if we just process it.
    
    raw_config = args.key_columns
    
    # Check if we have the complex syntax (semicolon or colon)
    if ':' in raw_config or ';' in raw_config:
        parts = raw_config.split(';')
        custom_default_set = False
        
        for part in parts:
            part = part.strip()
            if not part: continue
            
            if ':' in part:
                s_name, cols = part.split(':', 1)
                sheet_keys_config[s_name.strip()] = [c.strip() for c in cols.split(',')]
            else:
                # This part is a default list (e.g. "A,B" in "Sheet1:C;A,B")
                global_default_keys = [c.strip() for c in part.split(',')]
                custom_default_set = True
                
        # If user provided a string like "Sheet1:A" but no default part, global_default_keys keeps the initial default.
    else:
        # Simple comma-separated list, applies globally
        global_default_keys = [k.strip() for k in raw_config.split(',')]

    if not os.path.exists(input_file):
        print(f"Error: Input file not found at {input_file}")
        return

    print(f"Reading {input_file}...")
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return

    input_dir = os.path.dirname(input_file)

    for sheet_name in wb.sheetnames:
        output_filename = f"{sheet_name}.sql"
        output_path = os.path.join(input_dir, output_filename)
        
        print(f"Processing sheet: {sheet_name}")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows())

        # Determine keys for this sheet logic
        key_columns = global_default_keys
        action_col_idx = 0
        headers = []

        if rows:
            # Header is first row
            header_cells = rows[0]
            headers = [str(c.value).strip() if c.value else '' for c in header_cells]
            
            # Identify Action Column
            possible_action_headers = ['ACCION', 'ACTION', 'TIPO', 'OPERACION', 'MOVIMIENTO']
            
            # Try to find a header that matches
            found_action_col = False
            for i, h in enumerate(headers):
                if h.upper() in possible_action_headers:
                    action_col_idx = i
                    found_action_col = True
                    break
            
            print(f"  -> Stats: {len(rows)} rows found.")
            print(f"  -> Action Column: '{headers[action_col_idx]}' (Index {action_col_idx})")
            
            # Check for UPDATE action in the action column
            has_update = False
            for row in rows[1:]: # Skip header
                if len(row) > action_col_idx and row[action_col_idx].value:
                    val = str(row[action_col_idx].value).strip().upper()
                    if val == 'UPDATE':
                        has_update = True
                        break
            
            if has_update:
                if sheet_name in sheet_keys_config:
                    key_columns = sheet_keys_config[sheet_name]
                    print(f"  -> UPDATE detected. Using pre-configured keys: {key_columns}")
                else:
                    # Interactive mode
                    # Get all valid headers
                    available_cols = [h for h in headers if h]
                    
                    print(f"\n[!] UPDATE operations detected in '{sheet_name}'")
                    print(f"    Please select the Key Columns for the WHERE clause:")
                    
                    # Display columns in a list
                    for i, col in enumerate(available_cols, 1):
                        print(f"      {i}. {col}")

                    default_str = ", ".join(global_default_keys)
                    
                    while True:
                        user_input = input(f"    Enter column numbers separated by commas (e.g., 1,3) [Default: {default_str}]: ").strip()
                        
                        if not user_input:
                            key_columns = global_default_keys
                            print(f"    -> Using default keys: {key_columns}\n")
                            break
                        
                        try:
                            # Allow comma or space separation
                            selected_indices = [int(x.strip()) for x in user_input.replace(',', ' ').split() if x.strip()]
                            
                            # Validate indices
                            current_selection = []
                            invalid_indices = []
                            for idx in selected_indices:
                                if 1 <= idx <= len(available_cols):
                                    current_selection.append(available_cols[idx-1])
                                else:
                                    invalid_indices.append(idx)
                            
                            if invalid_indices:
                                print(f"    [!] Error: Invalid column numbers: {invalid_indices}")
                                continue
                                
                            if not current_selection:
                                print("    [!] Error: No columns selected.")
                                continue
                                
                            key_columns = current_selection
                            print(f"    -> Selected keys: {key_columns}\n")
                            break
                        except ValueError:
                            print("    [!] Error: Please enter valid numbers.")
        else:
            print("  -> Sheet is empty.")
        
        print(f"Generating {output_path}...")

        with open(output_path, 'w', encoding='utf-8') as sql_file:
            sql_file.write(f"BEGIN TRAN TRAN_{sheet_name}\n\n")
            sql_file.write("BEGIN TRY\n")
            
            if not rows:
                sql_file.write("-- Sheet is empty\n")
            else:
                # Header is first row
                header_cells = rows[0]
                # Headers are already loaded above
                # headers = [c.value for c in header_cells]
                
                # Map column name to index
                col_map = {h: i for i, h in enumerate(headers) if h}
                
                # Check if key columns exist
                has_keys = all(k in col_map for k in key_columns)
                
                # Iterate data rows
                generated_rows = 0
                for row_idx, row in enumerate(rows[1:], start=2):
                    # Action Column Check
                    if len(row) <= action_col_idx:
                        continue

                    action_cell = row[action_col_idx] 
                    action = str(action_cell.value).strip().upper() if action_cell.value else ''
                    
                    if action not in ('INSERT', 'UPDATE'):
                        continue
                    
                    generated_rows += 1

                    if action == 'INSERT':
                        # Generate INSERT for all columns (skipping the Action column)
                        
                        cols = []
                        vals = []
                        
                        for i, cell in enumerate(row):
                            if i == action_col_idx: continue # Skip Action column
                            if i >= len(headers) or not headers[i]: continue # Skip empty header cols
                            
                            cols.append(headers[i])
                            vals.append(format_value(cell.value))
                        
                        if cols:
                            col_str = ", ".join(cols)
                            val_str = ", ".join(vals)
                            sql_file.write(f"INSERT INTO {sheet_name} ({col_str}) VALUES ({val_str});\n")

                    elif action == 'UPDATE':
                        if not has_keys:
                            msg = f"-- ERROR: Missing key columns {key_columns} in sheet {sheet_name}\n"
                            sql_file.write(msg)
                            print(f"    [!] Error: {msg.strip()}")
                            continue

                        set_clauses = []
                        where_clauses = []
                        
                        # Build WHERE clause (always based on keys, regardless of style)
                        for key in key_columns:
                            idx = col_map[key]
                            val = format_value(row[idx].value)
                            where_clauses.append(f"{key} = {val}")

                        # Build SET clause (only styled cells)
                        for i, cell in enumerate(row):
                            if i == action_col_idx: continue # Skip Action column
                            if i >= len(headers) or not headers[i]: continue
                            
                            col_name = headers[i]
                            
                            # Skip if it is a key column (usually keys are not updated, but if styled maybe?)
                            # Let's assume keys are NOT updated even if styled, to be safe, 
                            # or allow update? Standard SQL updates non-keys.
                            if col_name in key_columns:
                                continue

                            if is_cell_styled(cell):
                                val = format_value(cell.value)
                                set_clauses.append(f"{col_name} = {val}")
                        
                        if set_clauses:
                            set_str = ", ".join(set_clauses)
                            where_str = " AND ".join(where_clauses)
                            sql_file.write(f"UPDATE {sheet_name} SET {set_str} WHERE {where_str};\n")
                
                if generated_rows == 0:
                    print(f"  -> Warning: No SQL statements were generated for '{sheet_name}'. Check Action column values.")
                    sql_file.write("-- Warning: No valid INSERT/UPDATE rows found.\n")

            sql_file.write(f"\nCOMMIT TRAN TRAN_{sheet_name};\n")
            sql_file.write("PRINT 'PROCESO EJECUTADO CORRECTAMENTE';\n")
            sql_file.write("END TRY\n")
            sql_file.write("BEGIN CATCH\n")
            sql_file.write("    SELECT 'LINEA ERROR - ' + CAST(ERROR_LINE() AS VARCHAR(5)) + ': ' + ERROR_MESSAGE();\n")
            sql_file.write(f"    ROLLBACK TRAN TRAN_{sheet_name};\n")
            sql_file.write("    PRINT 'OCURRIO UN ERROR EN EL PROCESO';\n")
            sql_file.write("END CATCH\n")

    print("All scripts generated successfully.")

if __name__ == "__main__":
    main()
